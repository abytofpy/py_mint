import json
from config import sets_to_reference, ROOT_DIR, languages_to_reference
from openpyxl import Workbook
from pm_collections import collections, look_for_card_in_collections
from pm_card import card
from pm_reference import build_reference, download_JSON
import inspect

#download_JSON(['KHM','M21'])
reference = (card_reference, name_to_uuid, number_to_uuid, uuid_to_number) = build_reference(sets_to_reference)            
### End of reference initialisation ####

# Tests begin here                                      
#decks_to_add = ['MysticIntellect_C19.json']
# test_collec = collections('test')
#parsed_cards = test_collec.from_parsed_source('additions_29012021', reference)
#parsed_cards = test_collec.from_parsed_source('deck_comm_legends', card_reference)
#test_collec.save('test.json')

# Loading collection
Rin_et_Seri_deck = collections('Rin_et_Seri')
print('Loading Rin et Seri ...')
parsed_cards1 = Rin_et_Seri_deck.from_parsed_source('Rin_et_Seri', reference)

Dalakos_deck = collections('Dalakos')
print('Loading Dalakos ...')
parsed_cards2 = Dalakos_deck.from_parsed_source('Dalakos', reference)

Atla_Palani_deck = collections('Atla_Palani')
print('Loading Atla_Palani ...')
parsed_cards3 = Atla_Palani_deck.from_parsed_source('Atla_Palani', reference)

Kaldheim_deck = collections('Kaldheim')
print('Loading Kaldheim ...')
parsed_cards4 = Kaldheim_deck.from_parsed_source('Kaldheim', reference)

Kykar_deck = collections('Kykar')
print('Loading Kykar ...')
parsed_cards5 = Kykar_deck.from_parsed_source('Kykar', reference)

Teshar_deck = collections('Teshar')
print('Loading Teshar ...')
parsed_cards6 = Teshar_deck.from_parsed_source('Teshar', reference)

Zaxara_deck = collections('Zaxara')
print('Loading Zaxara ...')
parsed_cards7 = Zaxara_deck.from_parsed_source('Zaxara', reference)

M2X_deck = collections('Additions_02012021')
print('Loading Additions_02012021 ...')
parsed_cards8 = M2X_deck.from_parsed_source('Additions_02012021', reference)

collections = [Rin_et_Seri_deck,
                Dalakos_deck,
                Atla_Palani_deck,
                Kaldheim_deck,
                Kykar_deck,
                Teshar_deck,
                Zaxara_deck,
                M2X_deck]

#print (Rin_et_Seri_deck[1])

#collection = get_inventory(collections, reference)


#def get_inventory(collections, reference):
    #"""
    #"""
inventory = []
for collection in collections:
    for card in collection.traces_cards(reference):
        inventory.append(card)
#    return(inventory)


print(len(inventory))
print (inventory[0])

# Test Output to excel
# workbook = Workbook()
# sheet = workbook.active
# for i,card in enumerate(parsed_cards):
    # sheet.cell(row=i+1, column=1).value = card[0]
    # sheet.cell(row=i+1, column=2).value = card[1]
# workbook.save(filename="Rin_et_Seri.xlsx")

#test_collec.list_cards(card_reference)

# test card movements with token creation :
# test_collec.save('test.json')
# card_to_move = define_card_from_set_and_number('TMP', '290', 'NM', 'EN', [])
# test_collec2 = collections('test2')
# test_collec.move_card_from_self_to_destination_replace_with_proxy(card_to_move, 'TMP', test_collec2) # Hack to be removed after
# test_collec.save('test.json')
# test_collec2.save('test2.json')

#Kykar_deck.save('Kykar.json')
#Teshar_deck.list_cards_edhrec_format(card_reference)

# print(" ------------------------------------------")
# print(" Test of research of cards in collections :")
# objects = dir()
# collections = [ eval(x) for x in objects if isinstance(eval(x), collections)] #inspect.isclass(x) == inspect.isclass(Teshar_deck) ]
# res = look_for_card_in_collections("Kykar, furie du vent", collections)
# if len(res) > 0 :
#     print ('.Found results:')
#     for element in res :
#         print (" - deck'{name}' : card from {setname}, condition {condition}, in {language}".format(name = element[0],
#                                                                                                 setname = element[1][0],
#                                                                                                 condition = element[1][2],
#                                                                                                 language = element[1][3]))
